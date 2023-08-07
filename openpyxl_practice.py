import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
print(wb.sheetnames)
ws1 = wb['Sheet1']
ws2 = wb.create_sheet('Sheet2')
ws3 = wb.create_sheet('Sheet3')
ws3.title = "Title_Upd"
print(f"After Adding two sheets and updating one of their name:{wb.sheetnames}")
values = ws1['A1:C2']
for row in values:
    for cell in row:
        print(cell.value)
"""
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)
<Cell Sheet1.A1>
<Cell Sheet1.B1>
<Cell Sheet1.C1>
<Cell Sheet1.A2>
<Cell Sheet1.B2>
<Cell Sheet1.C2>

for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)
<Cell Sheet1.A1>
<Cell Sheet1.A2>
<Cell Sheet1.B1>
<Cell Sheet1.B2>
<Cell Sheet1.C1>
<Cell Sheet1.C2>
"""""
