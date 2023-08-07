import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for i in range(2, sheet.max_row+1):
    curr_val = sheet.cell(i, 3).value
    updated_val = curr_val * 0.9
    corrected_cell = sheet.cell(i, 4)
    corrected_cell.value = updated_val
    values = Reference(sheet,min_row=2,max_row=4,min_col=4,max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
wb.save('Transaction_Updated.xlsx')

